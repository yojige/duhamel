# -*- coding: utf-8 -*-
import appex
import numpy as np
import math
import matplotlib.pyplot as plt
from matplotlib import gridspec
import openpyxl as px
import ui

acc = list()

def resp(sender):

    '''
    dt = 0.01
    min_period = 0.02
    max_period = 5.00
    bunkatu = 300
    damp = 0.05
    '''

    dt = float(sender.superview['textfield1'].text)
    min_period = float(sender.superview['textfield2'].text)
    max_period = float(sender.superview['textfield3'].text)
    bunkatu = int(sender.superview['textfield4'].text)
    damp = float(sender.superview['textfield5'].text)

    print('時間刻み　= ',dt)
    print('最小周期　= ',min_period)
    print('最大周期　= ',max_period)
    print('周期分割　= ',bunkatu)
    print('減衰　　　= ',damp)

    period = list()
    maxacc = list()

    dlta = ( math.log(max_period) - math.log(min_period) ) / float( bunkatu - 1 )
    for i in range( 1, bunkatu + 1 ):

        tmp = math.log(min_period) + dlta * float( i - 1 )
        period.append( math.exp(tmp) )	
        zmax = duhm(dt, damp, period[-1], acc)
        maxacc.append( zmax )	

    t = list()
    for i in range(0,len(acc)):
        t.append( float(i+1) * dt )
 
    # グラフ表示
    plt.figure(figsize=(8,12))
    plt.subplot(211)
    plt.plot(t, acc)
    plt.ylim(-rnd(abs_max(acc)), rnd(abs_max(acc)))
    plt.xlabel("Time(sec)")
    plt.ylabel("Acceleration$(cm/s^{2})$")

    plt.subplot(212)
    plt.plot(period, maxacc)
    plt.xscale('log')
    plt.xlabel("Period(sec)")
    plt.ylabel("Response Acceleration$(cm/s^{2})$")
    plt.grid(which='major',color='k',linestyle='-')
    plt.grid(which='minor',color='k',linestyle='--')
    plt.show()
    v.close()

def duhm(dt, h, T, a):
    w = 2.0 * math.pi / T
    w2 = w * w
    hw = h * w
    wd = w * math.sqrt( 1.0 - h * h )
    wdt = wd * dt
    e = math.exp( -hw * dt )
    cwdt = math.cos( wdt )
    swdt = math.sin( wdt )
    e11 = e * ( cwdt - hw * swdt / wd )
    e12 = -e * ( wd * wd + hw * hw ) * swdt / wd
    e21 = e * swdt / wd
    e22 = e * ( cwdt + hw * swdt / wd )
    ss = - hw * swdt - wd * cwdt
    cc = - hw * cwdt + wd * swdt
    s1 = ( e * ss + wd ) / w2
    c1 = ( e * cc + hw ) / w2
    s2 = ( e * dt * ss + hw * s1 + wd * c1 ) / w2
    c2 = ( e * dt * cc + hw * c1 - wd * s1 ) /w2
    s3 = dt * s1 - s2
    c3 = dt * c1 - c2
    g11 = ( -hw * s3 + wd * c3 ) / wdt
    g12 = ( -hw * s2 + wd * c2 ) / wdt
    g21 = s3 / wdt
    g22 = s2 / wdt
    dx = 0.0
    x = 0.0
    sa = 0.0
    sv = 0.0
    sd = 0.0
    accmax = 0.0
    for m in range(1, len(a)):
      dxf = dx
      xf = x
      ddym = a[ m ]
      ddyf = a[ m - 1 ]
      dx = e11 * dxf + e12 * xf + g11 * ddym + g12 * ddyf
      x = e21 * dxf + e22 * xf + g21 * ddym + g22 * ddyf
      ddx = 2.0 * hw * dx + w2 * x
      if abs(ddx) > abs(accmax):
          accmax = abs(ddx)

    return(accmax)


def rnd(v):
    a = math.log10(abs(v))
    if a < 0.0:
        n = int(a) - 1
    else:
        n = int(a)
    b = 10.0**float(n)
    m = int( abs(v) / b ) + 1
    r = float(m) * b
    return(r)


def abs_max(v):
    mymax = max(v)
    mymin = min(v)
    if abs(mymax) > abs(mymin):
        r = abs(mymax)
    else:
        r = abs(mymin)
    return(r)


myf = appex.get_file_path()
wb = px.load_workbook(myf, data_only=True)
#wb = px.load_workbook('test.xlsx', data_only=True)
ws = wb.worksheets[0]
for cell in ws['A']:
    acc.append(cell.value)

v = ui.load_view()
v.present('fullscreen')
